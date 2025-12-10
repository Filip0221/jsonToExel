from pathlib import Path
import json
import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

def jsonToExel(input_dir, output_file):
    # Używamy Pandas ExcelWritera z silnikiem openpyxl, żeby móc później użyć zaawansowanego formatowania
    writer = pd.ExcelWriter(output_file, engine="openpyxl")
    
    # Przechodzimy przez każdy plik .json w wybranym folderze
    for json_file in input_dir.glob("*.json"):
        # print(f"Przetwarzam plik: {json_file.name}") # do sprawdzania czy łapie wszystkie pliki
        
        # Wczytanie całej zawartości JSON
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Przygotowujemy dane do zapisania. Każdy główny element JSON będzie osobną małą tabelą
        tables = {}
        for bulb_key, bulb_data in data.items():
            df = pd.DataFrame([bulb_data])

            # żeby były osobne kolumny w exelu dla okien pradowych bez teog bedzą okna w jednej komurce
            if "thresholds" in df.columns:
                thresholds = df["thresholds"].iloc[0]
                for i, val in enumerate(thresholds):
                    df[f"threshold_{i+1}"] = val
                df = df.drop(columns=["thresholds"])

            tables[bulb_key] = df

        sheet_name = json_file.stem
        
        # Musimy najpierw stworzyć pusty arkusz przez Pandas, aby móc go pobrać jako obiekt openpyxl
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name] # obiekt do formatowania
        
        start_row = 0 # Wiersz startowy do zapisu danych w arkuszu

        # Zapisujemy każdą małą tabelę po kolei w nowym arkuszu
        for name, df in tables.items():
            
            # Najpierw wstawiamy i formatujemy tytuł tabeli
            num_cols = max(1, len(df.columns))
            title_row_excel = start_row + 1 #openpyxl liczy wiersze od 1 !!!!!!!!!!!!

            cell = ws.cell(row=title_row_excel, column=1, value=name)
            
            # Scalamy komórki na szerokość tabeli i formatujemy
            ws.merge_cells(
                start_row=title_row_excel,
                start_column=1,
                end_row=title_row_excel,
                end_column=num_cols
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
            cell.font = Font(bold=True)

            start_row += 1 # Przesuwamy się w dół, żeby dane trafiły pod tytuł

            # Zapisujemy dane i nagłówki tabeli Pandas
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

            # Dopasowujemy automatycznie szerokość kolumn
            num_cols = len(df.columns)
            for col_idx in range(1, num_cols + 1):
                col_letter = get_column_letter(col_idx)

                # Obliczamy maksymalną długość tekstu w kolumnie
                max_length = len(df.columns[col_idx-1]) 
                for row_idx in range(start_row + 1, start_row + len(df) + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))

                ws.column_dimensions[col_letter].width = max_length + 2 # +2 na odstęp

            # Obliczamy nowy wiersz startowy, zostawiając jedną pustą linię przerwy.
            rows_written = 1 + len(df)
            start_row = start_row + rows_written + 1

    # Zapisujemy i zamykamy cały plik Excela
    writer.close()
    # print("Gotowe")